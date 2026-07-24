from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import time
import uuid
from pathlib import Path
from typing import Any, Iterable

import fitz
import httpx
from openpyxl import load_workbook
from qdrant_client import QdrantClient
from qdrant_client.http import models

try:
    from .archive import archive_pdfs
except ImportError:  # Script execution from the image entrypoint.
    from archive import archive_pdfs


DATA_DIR = Path(os.getenv("COURSE_DATA_DIR", "/data/raw"))
REPORT_DIR = Path(os.getenv("REPORT_DIR", "/data/reports"))
QDRANT_URL = os.getenv("QDRANT_URL", "http://qdrant:6333")
QDRANT_API_KEY = os.getenv("QDRANT_API_KEY") or None
COLLECTION = os.getenv("QDRANT_COLLECTION", "storage-course-v1")
EMBEDDING_URL = os.getenv("EMBEDDING_URL", "http://model-gateway:8080/v1/embeddings")
EMBEDDING_MODEL = os.getenv("EMBEDDING_MODEL", "mock-embedding")
MOCK_EMBEDDINGS = os.getenv("MOCK_EMBEDDINGS", "false").lower() in {"1", "true", "yes", "on"}
EMBEDDING_DIMENSION = int(os.getenv("MOCK_EMBEDDING_DIMENSION", "64"))

CHAPTERS = {
    1: "第1章 概述",
    2: "第2章 电力系统与储能技术的应用",
    3: "第3章 电力储能系统的组成及工作原理",
    4: "第4章 电力储能系统的规划配置",
    5: "第5章 电力储能系统的接入与运行控制",
    6: "第6章 电力储能系统的性能检测与评估",
}


def stable_vector(text: str) -> list[float]:
    values: list[float] = []
    counter = 0
    while len(values) < EMBEDDING_DIMENSION:
        digest = hashlib.sha256(f"{text}:{counter}".encode()).digest()
        values.extend((item / 127.5) - 1.0 for item in digest)
        counter += 1
    return values[:EMBEDDING_DIMENSION]


def chapter_number(path: Path) -> int | None:
    text = str(path)
    match = re.search(r"(?:第)?([1-6])(?:章|[._-])", text)
    return int(match.group(1)) if match else None


def input_pdfs() -> tuple[list[Path], dict[str, str]]:
    pdfs = sorted(DATA_DIR.rglob("*.pdf"))
    if pdfs:
        names = {pdf.name: pdf.name for pdf in pdfs}
        manifest_path = DATA_DIR / "normalized" / "manifest.json"
        if manifest_path.exists():
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                names.update(
                    {
                        str(item["normalized_file"]): str(item["source_file"])
                        for item in manifest.get("files", [])
                        if item.get("normalized_file") and item.get("source_file")
                    }
                )
            except (OSError, json.JSONDecodeError, TypeError):
                pass
        return pdfs, names
    root, source_names = archive_pdfs(DATA_DIR)
    return sorted(root.glob("*.pdf")), source_names


def load_knowledge_points() -> list[dict[str, Any]]:
    files = list(DATA_DIR.glob("*.xlsx"))
    if not files:
        return []
    points: list[dict[str, Any]] = []
    workbook = load_workbook(files[0], read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            joined = " ".join(values)
            match = re.search(r"(\d+\.\d+)\s*([^ ]+.*)", joined)
            if match and not any(item["id"] == match.group(1) for item in points):
                points.append({"id": match.group(1), "name": match.group(2).strip()})
    return points


def resolve_knowledge_point(text: str, points: list[dict[str, Any]]) -> str | None:
    for point in points:
        if point["id"] in text or point["name"] in text:
            return f"{point['id']} {point['name']}"
    return None


def chunk_text(text: str, size: int = 900, overlap: int = 120) -> Iterable[str]:
    cleaned = re.sub(r"\s+", " ", text).strip()
    if not cleaned:
        return
    start = 0
    while start < len(cleaned):
        end = min(start + size, len(cleaned))
        yield cleaned[start:end]
        if end >= len(cleaned):
            break
        start = end - overlap


def extract_documents(points: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    pdfs, source_names = input_pdfs()
    chunks: list[dict[str, Any]] = []
    report: dict[str, Any] = {"pdf_count": len(pdfs), "files": [], "errors": [], "unmapped": []}
    for pdf in pdfs:
        display_source = source_names.get(pdf.name, pdf.name)
        file_entry: dict[str, Any] = {"file": display_source, "pages": 0, "text_pages": 0, "chunks": 0}
        try:
            document = fitz.open(pdf)
            file_entry["pages"] = len(document)
            chapter_id = chapter_number(pdf)
            chapter = CHAPTERS.get(chapter_id, "未分类")
            knowledge_point = resolve_knowledge_point(f"{pdf.name} {display_source}", points)
            if chapter_id is None:
                report["unmapped"].append(file_entry["file"])
            for page_index, page in enumerate(document):
                text = page.get_text("text").strip()
                if not text:
                    continue
                file_entry["text_pages"] += 1
                for chunk_index, content in enumerate(chunk_text(text)):
                    point_id = knowledge_point or "未映射"
                    identity = f"{display_source}:{page_index + 1}:{chunk_index}"
                    # Preserve a machine-checkable citation in the text sent
                    # to any downstream knowledge-base experiment. The
                    # primary Xingchen path validates this marker against its
                    # local manifest before showing it to a user.
                    citation = f"[来源文件：{display_source}；章节：{chapter}；页码：{page_index + 1}]\n"
                    chunks.append(
                        {
                            "id": str(uuid.uuid5(uuid.NAMESPACE_URL, identity)),
                            "text": citation + content,
                            "metadata": {
                                "source_file": display_source,
                                "chapter": chapter,
                                "chapter_id": chapter_id,
                                "knowledge_point": point_id,
                                "page": page_index + 1,
                                "chunk_index": chunk_index,
                            },
                        }
                    )
                    file_entry["chunks"] += 1
            document.close()
            report["files"].append(file_entry)
        except Exception as exc:  # Keep one bad file from aborting the batch.
            report["errors"].append({"file": file_entry["file"], "error": str(exc)})
    report["chunk_count"] = len(chunks)
    report["text_page_count"] = sum(item["text_pages"] for item in report["files"])
    report["page_count"] = sum(item["pages"] for item in report["files"])
    return chunks, report


def embed(texts: list[str]) -> list[list[float]]:
    if MOCK_EMBEDDINGS:
        return [stable_vector(text) for text in texts]
    response = httpx.post(
        EMBEDDING_URL,
        headers={"content-type": "application/json"},
        json={"model": EMBEDDING_MODEL, "input": texts},
        timeout=90,
    )
    response.raise_for_status()
    payload = response.json()
    return [item["embedding"] for item in sorted(payload["data"], key=lambda item: item["index"])]


def upsert(chunks: list[dict[str, Any]]) -> dict[str, Any]:
    if not chunks:
        return {"collection": COLLECTION, "points": 0}
    client = QdrantClient(url=QDRANT_URL, api_key=QDRANT_API_KEY, timeout=10)
    deadline = time.time() + 120
    last_error: Exception | None = None
    while time.time() < deadline:
        try:
            client.get_collections()
            break
        except Exception as exc:
            last_error = exc
            time.sleep(2)
    else:
        raise RuntimeError(f"Qdrant did not become ready: {last_error}")
    vector_size = EMBEDDING_DIMENSION if MOCK_EMBEDDINGS else len(embed([chunks[0]["text"]])[0])
    if client.collection_exists(COLLECTION):
        client.delete_collection(COLLECTION)
    client.create_collection(
        collection_name=COLLECTION,
        vectors_config=models.VectorParams(size=vector_size, distance=models.Distance.COSINE),
    )
    batch_size = 32
    for start in range(0, len(chunks), batch_size):
        batch = chunks[start : start + batch_size]
        vectors = embed([item["text"] for item in batch])
        client.upsert(
            collection_name=COLLECTION,
            points=[
                models.PointStruct(
                    id=item["id"],
                    vector=vector,
                    # Keep Flowise's default content/metadata contract while
                    # retaining flat fields for direct Qdrant filtering.
                    payload={
                        "content": item["text"],
                        "text": item["text"],
                        "metadata": item["metadata"],
                        **item["metadata"],
                    },
                )
                for item, vector in zip(batch, vectors)
            ],
            wait=True,
        )
    return {"collection": COLLECTION, "points": len(chunks), "vector_size": vector_size}


def main() -> int:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    points = load_knowledge_points()
    chunks, report = extract_documents(points)
    report["knowledge_point_count"] = len(points)
    try:
        report["qdrant"] = upsert(chunks)
    except Exception as exc:
        report["errors"].append({"stage": "qdrant", "error": str(exc)})
    report_path = REPORT_DIR / "ingestion-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 1 if report["errors"] else 0


if __name__ == "__main__":
    sys.exit(main())
